#!/usr/bin/env python3
"""Build an Excel workbook with raw data and 12 charts comparing V2 (Simple DPLL) and MiniSat.

Outputs:
    test_results/solver_metrics_workbook.xlsx
"""

from __future__ import annotations

import csv
import re
import sys
import statistics
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / ".pip_local"))

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference, ScatterChart, Series
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.chart.trendline import Trendline
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_PATH = ROOT / "test_results" / "solver_metrics_workbook.xlsx"
RAW_CSV = ROOT / "test_results" / "all_solver_metric_records.csv"

VAR_BUCKETS = [50, 75, 125, 200]

V2_LABEL = "Our Simple DPLL (V2)"
MINISAT_LABELS = ("MiniSat (RSS rerun)", "MiniSat (original run)")
MINISAT_DISPLAY = "MiniSat"


def load_rows():
    rows = []
    with RAW_CSV.open(newline="", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            try:
                vars_ = int(r["Variables"]) if r["Variables"] else None
            except ValueError:
                vars_ = None
            if vars_ not in VAR_BUCKETS:
                continue
            try:
                wall = float(r["Wall Time [ms]"]) if r["Wall Time [ms]"] else None
                cpu = float(r["CPU Time [ms]"]) if r["CPU Time [ms]"] else None
            except ValueError:
                continue
            mem_kb = None
            if r["Peak Memory [KB]"]:
                try:
                    mem_kb = float(r["Peak Memory [KB]"])
                except ValueError:
                    pass
            solver_raw = r["Solver"]
            if solver_raw == V2_LABEL:
                solver = V2_LABEL
            elif solver_raw in MINISAT_LABELS:
                solver = MINISAT_DISPLAY
            else:
                continue
            rows.append({
                "vars": vars_,
                "clauses": int(r["Clauses"]) if r["Clauses"] else None,
                "solver": solver,
                "solver_raw": solver_raw,
                "problem_type": r["Problem Type"] or "",
                "instance": r["Instance"],
                "result": r["Result"],
                "wall_ms": wall,
                "cpu_ms": cpu,
                "mem_kb": mem_kb,
            })
    return rows


HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FONT = Font(bold=True, size=12)


def write_header(ws, row, columns):
    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=row, column=col_idx, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def autosize(ws, max_col):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        max_len = 8
        for cell in ws[letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 32)


def stats(values):
    vals = [v for v in values if v is not None]
    if not vals:
        return {"avg": None, "min": None, "max": None, "median": None, "n": 0}
    return {
        "avg": sum(vals) / len(vals),
        "min": min(vals),
        "max": max(vals),
        "median": statistics.median(vals),
        "n": len(vals),
    }


def build_summary_sheet(wb, rows):
    ws = wb.create_sheet("Summary by Group")
    columns = [
        "Solver", "Problem Type", "Variables", "Clauses", "Instances",
        "Avg Wall [ms]", "Min Wall [ms]", "Max Wall [ms]", "Median Wall [ms]",
        "Avg CPU [ms]", "Min CPU [ms]", "Max CPU [ms]", "Median CPU [ms]",
        "Avg Mem [KB]", "Min Mem [KB]", "Max Mem [KB]",
    ]
    write_header(ws, 1, columns)
    out_row = 2
    for solver in (V2_LABEL, MINISAT_DISPLAY):
        for ptype in ("SAT", "UNSAT"):
            for vars_ in VAR_BUCKETS:
                bucket = [r for r in rows if r["solver"] == solver and r["problem_type"] == ptype and r["vars"] == vars_]
                if not bucket:
                    continue
                clauses = bucket[0]["clauses"]
                wall = stats(r["wall_ms"] for r in bucket)
                cpu = stats(r["cpu_ms"] for r in bucket)
                mem = stats(r["mem_kb"] for r in bucket)
                values = [
                    solver, ptype, vars_, clauses, wall["n"],
                    wall["avg"], wall["min"], wall["max"], wall["median"],
                    cpu["avg"], cpu["min"], cpu["max"], cpu["median"],
                    mem["avg"], mem["min"], mem["max"],
                ]
                for col_idx, val in enumerate(values, 1):
                    ws.cell(row=out_row, column=col_idx, value=val)
                out_row += 1
    autosize(ws, len(columns))
    ws.freeze_panes = "A2"
    return ws


def build_raw_sheet(wb, rows):
    ws = wb.create_sheet("Raw Data")
    columns = [
        "Variables", "Clauses", "Solver", "Solver Source", "Problem Type",
        "Instance", "Result",
        "Wall Time [ms]", "CPU Time [ms]", "Peak Memory [KB]",
    ]
    write_header(ws, 1, columns)
    for idx, r in enumerate(rows, 2):
        ws.cell(row=idx, column=1, value=r["vars"])
        ws.cell(row=idx, column=2, value=r["clauses"])
        ws.cell(row=idx, column=3, value=r["solver"])
        ws.cell(row=idx, column=4, value=r["solver_raw"])
        ws.cell(row=idx, column=5, value=r["problem_type"])
        ws.cell(row=idx, column=6, value=r["instance"])
        ws.cell(row=idx, column=7, value=r["result"])
        ws.cell(row=idx, column=8, value=r["wall_ms"])
        ws.cell(row=idx, column=9, value=r["cpu_ms"])
        ws.cell(row=idx, column=10, value=r["mem_kb"])
    autosize(ws, len(columns))
    ws.freeze_panes = "A2"
    return ws


def add_average_chart(ws, title, table_top, n_cols, anchor):
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.grouping = "clustered"
    chart.title = title
    chart.x_axis.title = "Variables"
    chart.y_axis.title = "Time [ms] (log scale)"
    chart.height = 10
    chart.width = 18
    chart.y_axis.scaling.logBase = 10
    data = Reference(ws, min_col=2, min_row=table_top, max_col=n_cols, max_row=table_top + len(VAR_BUCKETS))
    cats = Reference(ws, min_col=1, min_row=table_top + 1, max_row=table_top + len(VAR_BUCKETS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def write_average_table(ws, title, start_row, headers, data_rows):
    ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
    write_header(ws, start_row + 1, headers)
    for i, drow in enumerate(data_rows, 1):
        for j, val in enumerate(drow, 1):
            ws.cell(row=start_row + 1 + i, column=j, value=val)
    return start_row + 1


def build_avg_sheet(wb, rows, solver, sheet_name):
    ws = wb.create_sheet(sheet_name)
    walls = [["Variables"] + ["SAT", "UNSAT"]]
    cpus = [["Variables"] + ["SAT", "UNSAT"]]
    mems = [["Variables"] + ["SAT", "UNSAT"]]
    for vars_ in VAR_BUCKETS:
        sat_bucket = [r for r in rows if r["solver"] == solver and r["problem_type"] == "SAT" and r["vars"] == vars_]
        uns_bucket = [r for r in rows if r["solver"] == solver and r["problem_type"] == "UNSAT" and r["vars"] == vars_]
        walls.append([vars_, stats(r["wall_ms"] for r in sat_bucket)["avg"], stats(r["wall_ms"] for r in uns_bucket)["avg"]])
        cpus.append([vars_, stats(r["cpu_ms"] for r in sat_bucket)["avg"], stats(r["cpu_ms"] for r in uns_bucket)["avg"]])
        mems.append([vars_, stats(r["mem_kb"] for r in sat_bucket)["avg"], stats(r["mem_kb"] for r in uns_bucket)["avg"]])

    cur = 1
    headers = walls[0]
    cur = write_average_table(ws, f"{solver} – Average Wall Time [ms]", cur, headers, walls[1:])
    add_average_chart(ws, f"{solver}: Avg Wall Time vs Variables", cur, len(headers), f"E{cur}")
    cur += len(VAR_BUCKETS) + 22

    cur = write_average_table(ws, f"{solver} – Average CPU Time [ms]", cur, headers, cpus[1:])
    add_average_chart(ws, f"{solver}: Avg CPU Time vs Variables", cur, len(headers), f"E{cur}")
    cur += len(VAR_BUCKETS) + 22

    cur = write_average_table(ws, f"{solver} – Average Peak Memory [KB]", cur, headers, mems[1:])
    add_memory_chart(ws, f"{solver}: Avg Peak Memory vs Variables", cur, len(headers), f"E{cur}")
    autosize(ws, 6)


def add_memory_chart(ws, title, table_top, n_cols, anchor):
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.grouping = "clustered"
    chart.title = title
    chart.x_axis.title = "Variables"
    chart.y_axis.title = "Peak Memory [KB]"
    chart.height = 10
    chart.width = 18
    data = Reference(ws, min_col=2, min_row=table_top, max_col=n_cols, max_row=table_top + len(VAR_BUCKETS))
    cats = Reference(ws, min_col=1, min_row=table_top + 1, max_row=table_top + len(VAR_BUCKETS))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, anchor)


def build_scatter_sheet(wb, rows, solver, sheet_name):
    """Per-instance scatter: x=vars (with jitter), y=wall ms (log)."""
    ws = wb.create_sheet(sheet_name)
    headers = ["Variables (jittered)", "SAT Wall [ms]", "UNSAT Wall [ms]"]
    write_header(ws, 1, headers)

    bucket_rows = {v: {"SAT": [], "UNSAT": []} for v in VAR_BUCKETS}
    for r in rows:
        if r["solver"] != solver or r["vars"] not in VAR_BUCKETS or r["problem_type"] not in ("SAT", "UNSAT"):
            continue
        if r["wall_ms"] is None or r["wall_ms"] <= 0:
            continue
        bucket_rows[r["vars"]][r["problem_type"]].append(r["wall_ms"])

    out_row = 2
    for vars_ in VAR_BUCKETS:
        sat_vals = bucket_rows[vars_]["SAT"]
        uns_vals = bucket_rows[vars_]["UNSAT"]
        for i, v in enumerate(sat_vals):
            jitter = -1.5 + (i % 7) * 0.4
            ws.cell(row=out_row, column=1, value=vars_ + jitter)
            ws.cell(row=out_row, column=2, value=v)
            out_row += 1
        for i, v in enumerate(uns_vals):
            jitter = 1.5 + (i % 7) * 0.4
            ws.cell(row=out_row, column=1, value=vars_ + jitter)
            ws.cell(row=out_row, column=3, value=v)
            out_row += 1

    last_row = out_row - 1
    chart = ScatterChart()
    chart.title = f"{solver}: Per-Instance Wall Time (log scale)"
    chart.style = 13
    chart.x_axis.title = "Variables"
    chart.y_axis.title = "Wall Time [ms] (log)"
    chart.height = 14
    chart.width = 24
    chart.y_axis.scaling.logBase = 10

    xref = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    sat_ref = Reference(ws, min_col=2, min_row=2, max_row=last_row)
    uns_ref = Reference(ws, min_col=3, min_row=2, max_row=last_row)

    sat_series = Series(sat_ref, xref, title="SAT")
    uns_series = Series(uns_ref, xref, title="UNSAT")
    sat_series.graphicalProperties = None
    chart.series.append(sat_series)
    chart.series.append(uns_series)

    for s in chart.series:
        s.spPr = None
    ws.add_chart(chart, "E2")
    autosize(ws, 3)


def build_compare_sheet(wb, rows, sheet_name):
    """Aggregate comparison sheet: 4 charts comparing V2 vs MiniSat for SAT & UNSAT."""
    ws = wb.create_sheet(sheet_name)

    def averages_table(title, start_row, ptype, key):
        headers = ["Variables", V2_LABEL, MINISAT_DISPLAY]
        ws.cell(row=start_row, column=1, value=title).font = SECTION_FONT
        write_header(ws, start_row + 1, headers)
        for i, vars_ in enumerate(VAR_BUCKETS, 1):
            v2 = stats(r[key] for r in rows if r["solver"] == V2_LABEL and r["problem_type"] == ptype and r["vars"] == vars_)
            ms = stats(r[key] for r in rows if r["solver"] == MINISAT_DISPLAY and r["problem_type"] == ptype and r["vars"] == vars_)
            ws.cell(row=start_row + 1 + i, column=1, value=vars_)
            ws.cell(row=start_row + 1 + i, column=2, value=v2["avg"])
            ws.cell(row=start_row + 1 + i, column=3, value=ms["avg"])
        return start_row + 1

    def add_log_bar(title, table_top, anchor, y_axis_label):
        chart = BarChart()
        chart.type = "col"
        chart.style = 12
        chart.grouping = "clustered"
        chart.title = title
        chart.x_axis.title = "Variables"
        chart.y_axis.title = y_axis_label
        chart.height = 10
        chart.width = 20
        chart.y_axis.scaling.logBase = 10
        data = Reference(ws, min_col=2, min_row=table_top, max_col=3, max_row=table_top + len(VAR_BUCKETS))
        cats = Reference(ws, min_col=1, min_row=table_top + 1, max_row=table_top + len(VAR_BUCKETS))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, anchor)

    cur = 1
    cur = averages_table("Avg Wall Time [ms] – SAT", cur, "SAT", "wall_ms")
    add_log_bar("Comparison: Avg Wall Time SAT (log)", cur, f"E{cur}", "Avg Wall [ms] (log)")
    cur += len(VAR_BUCKETS) + 22

    cur = averages_table("Avg Wall Time [ms] – UNSAT", cur, "UNSAT", "wall_ms")
    add_log_bar("Comparison: Avg Wall Time UNSAT (log)", cur, f"E{cur}", "Avg Wall [ms] (log)")
    cur += len(VAR_BUCKETS) + 22

    cur = averages_table("Avg CPU Time [ms] – SAT", cur, "SAT", "cpu_ms")
    add_log_bar("Comparison: Avg CPU Time SAT (log)", cur, f"E{cur}", "Avg CPU [ms] (log)")
    cur += len(VAR_BUCKETS) + 22

    cur = averages_table("Avg CPU Time [ms] – UNSAT", cur, "UNSAT", "cpu_ms")
    add_log_bar("Comparison: Avg CPU Time UNSAT (log)", cur, f"E{cur}", "Avg CPU [ms] (log)")
    autosize(ws, 4)


def build_compare_scatter_sheet(wb, rows, sheet_name):
    """Per-instance scatter comparison: V2 vs MiniSat overlaid for SAT and UNSAT."""
    ws = wb.create_sheet(sheet_name)
    headers = ["Variables (jittered)", "V2 SAT", "V2 UNSAT", "MiniSat SAT", "MiniSat UNSAT"]
    write_header(ws, 1, headers)

    out_row = 2
    column_for = {
        (V2_LABEL, "SAT"): 2,
        (V2_LABEL, "UNSAT"): 3,
        (MINISAT_DISPLAY, "SAT"): 4,
        (MINISAT_DISPLAY, "UNSAT"): 5,
    }
    jitter_for = {
        (V2_LABEL, "SAT"): -2.5,
        (MINISAT_DISPLAY, "SAT"): -1.0,
        (V2_LABEL, "UNSAT"): 1.0,
        (MINISAT_DISPLAY, "UNSAT"): 2.5,
    }

    for vars_ in VAR_BUCKETS:
        for solver in (V2_LABEL, MINISAT_DISPLAY):
            for ptype in ("SAT", "UNSAT"):
                vals = [r["wall_ms"] for r in rows if r["solver"] == solver and r["problem_type"] == ptype and r["vars"] == vars_ and r["wall_ms"] and r["wall_ms"] > 0]
                base_jitter = jitter_for[(solver, ptype)]
                for i, v in enumerate(vals):
                    ws.cell(row=out_row, column=1, value=vars_ + base_jitter + (i % 5) * 0.15)
                    ws.cell(row=out_row, column=column_for[(solver, ptype)], value=v)
                    out_row += 1

    last_row = out_row - 1
    chart = ScatterChart()
    chart.title = "V2 vs MiniSat: Per-Instance Wall Time (log scale)"
    chart.style = 13
    chart.x_axis.title = "Variables"
    chart.y_axis.title = "Wall Time [ms] (log)"
    chart.height = 16
    chart.width = 26
    chart.y_axis.scaling.logBase = 10
    xref = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    for col in range(2, 6):
        yref = Reference(ws, min_col=col, min_row=2, max_row=last_row)
        title = ws.cell(row=1, column=col).value
        chart.series.append(Series(yref, xref, title=title))
    for s in chart.series:
        s.spPr = None
    ws.add_chart(chart, "G2")
    autosize(ws, 5)


def build_workbook():
    rows = load_rows()
    wb = Workbook()
    cover = wb.active
    cover.title = "README"
    cover.append(["Solver Metrics Workbook"])
    cover["A1"].font = Font(bold=True, size=14)
    cover.append([])
    cover.append([
        "This workbook summarizes Wall, CPU, and Memory metrics for two SAT solvers:",
    ])
    cover.append([f"  - {V2_LABEL}"])
    cover.append([f"  - {MINISAT_DISPLAY} (combined original and RSS rerun)"])
    cover.append([])
    cover.append(["Sheets:"])
    cover.append(["  Summary by Group: aggregate avg/min/max/median per solver-problem-vars bucket"])
    cover.append(["  Raw Data: every parsed instance row used downstream"])
    cover.append(["  V2 Averages: 3 charts (avg wall, avg cpu, avg memory) for our DPLL"])
    cover.append(["  MiniSat Averages: 3 charts for MiniSat"])
    cover.append(["  V2 Scatter: per-instance wall-time scatter (log Y) for V2"])
    cover.append(["  MiniSat Scatter: per-instance wall-time scatter (log Y) for MiniSat"])
    cover.append(["  Comparison Averages: 4 grouped bar charts comparing V2 vs MiniSat"])
    cover.append(["  Comparison Scatter: per-instance overlaid scatter (4 series in one chart)"])
    cover.append([])
    cover.append([f"Total raw rows considered (50/75/125/200 vars only): {len(rows)}"])
    cover.column_dimensions["A"].width = 90

    build_summary_sheet(wb, rows)
    build_raw_sheet(wb, rows)
    build_avg_sheet(wb, rows, V2_LABEL, "V2 Averages")
    build_avg_sheet(wb, rows, MINISAT_DISPLAY, "MiniSat Averages")
    build_scatter_sheet(wb, rows, V2_LABEL, "V2 Scatter")
    build_scatter_sheet(wb, rows, MINISAT_DISPLAY, "MiniSat Scatter")
    build_compare_sheet(wb, rows, "Comparison Averages")
    build_compare_scatter_sheet(wb, rows, "Comparison Scatter")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    print(f"rows: {len(rows)}")


if __name__ == "__main__":
    build_workbook()
